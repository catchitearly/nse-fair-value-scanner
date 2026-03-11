"""
report_generator.py
--------------------
Generates Excel report + console summary from valuation results.

Sheets:
  📊 Summary          — all stocks sorted by verdict
  ⭐ Strong Buy       — per-verdict sheet
  ✅ Buy
  🟡 Hold
  🔴 Avoid
  🚫 Strong Avoid
  📈 Monthly Returns  — backtest: return at 30,60,...,365 days per verdict
  📈 Backtest Accuracy— backtest: accuracy stats per verdict
  ℹ️ Metadata         — scan parameters
"""

import os
import logging
from datetime import datetime

import numpy as np
import pandas as pd
from tabulate import tabulate

logger = logging.getLogger(__name__)

VERDICT_ORDER = [
    "⭐ Strong Buy",
    "✅ Buy",
    "🟡 Hold",
    "🔴 Avoid",
    "🚫 Strong Avoid",
    "Insufficient Data",
]

RETURN_INTERVALS = [30, 60, 90, 120, 150, 180, 210, 240, 270, 300, 330, 365]

DISPLAY_COLS = [
    "ticker", "name", "sector", "scan_date", "mode",
    "price", "market_cap_cr",
    "fair_value_bear", "fair_value_base", "fair_value_bull",
    "margin_of_safety_pct", "quality_score", "verdict",
    "pe_ttm", "pe_fwd", "pb", "peg", "ev_ebitda",
    "roe", "roce", "net_margin", "op_margin",
    "revenue_growth", "earnings_growth",
    "debt_equity", "current_ratio",
    "div_yield", "beta",
    "week52_low", "week52_high", "pct_from_52w_high", "pct_from_52w_low",
    "rv", "gv", "dcf", "av", "models_used",
    "analyst_rec", "num_analysts", "target_mean",
]

# Monthly return columns appended dynamically
RETURN_COLS      = [f"return_{d}d" for d in RETURN_INTERVALS]
PRICE_FWD_COLS   = [f"price_{d}d"  for d in RETURN_INTERVALS]

PCT_COLS = [
    "roe", "roa", "roce", "net_margin", "op_margin", "gross_margin",
    "revenue_growth", "earnings_growth", "div_yield", "payout_ratio"
]


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(df):
    print("\n" + "═" * 95)
    print(f"  NSE FAIR VALUE SCANNER  |  {datetime.now().strftime('%d %b %Y  %H:%M')}")
    mode = df["mode"].iloc[0] if "mode" in df.columns and len(df) else "live"
    sd   = df["scan_date"].iloc[0] if "scan_date" in df.columns and len(df) else ""
    print(f"  Mode: {mode.upper()}  |  Scan date: {sd}  |  Stocks: {len(df)}")
    print("═" * 95)

    for verdict in VERDICT_ORDER:
        subset = df[df["verdict"] == verdict].copy()
        if subset.empty:
            continue
        subset = subset.sort_values("margin_of_safety_pct", ascending=False)

        print(f"\n{'─'*95}")
        print(f"  {verdict}  ({len(subset)} stocks)")
        print(f"{'─'*95}")

        rows = []
        for _, r in subset.iterrows():
            roe_pct = r.get("roe")
            roe_str = f"{roe_pct*100:.1f}%" if pd.notna(roe_pct) else ""
            rows.append([
                r.get("ticker", ""),
                (r.get("name", "") or "")[:26],
                (r.get("sector", "") or "")[:16],
                f"{r.get('price', ''):.2f}"          if pd.notna(r.get("price"))           else "",
                f"{r.get('market_cap_cr', ''):.0f}"  if pd.notna(r.get("market_cap_cr"))   else "",
                f"{r.get('fair_value_base', ''):.2f}" if pd.notna(r.get("fair_value_base")) else "",
                f"{r.get('margin_of_safety_pct', ''):.1f}%" if pd.notna(r.get("margin_of_safety_pct")) else "",
                f"{r.get('quality_score', ''):.0f}"  if pd.notna(r.get("quality_score"))   else "",
                f"{r.get('pe_fwd', ''):.1f}"         if pd.notna(r.get("pe_fwd"))          else "",
                roe_str,
            ])

        print(tabulate(rows, tablefmt="simple", headers=[
            "Ticker", "Name", "Sector", "Price", "MCap cr",
            "FairVal", "MoS%", "Quality", "FwdPE", "ROE%"
        ]))

    print("\n" + "═" * 95)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

def save_excel(df, output_dir="output", scan_date=None, is_backtest=False):
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    filepath = os.path.join(output_dir, f"nse_fair_value_{date_str}.xlsx")

    # ── Prepare main dataframe ─────────────────────────────────────────────────
    avail_display = [c for c in DISPLAY_COLS if c in df.columns]
    avail_ret     = [c for c in RETURN_COLS  if c in df.columns]
    all_cols      = avail_display + avail_ret
    df_out        = df[all_cols].copy()

    # Convert ratio cols to %
    for col in PCT_COLS:
        if col in df_out.columns:
            df_out[col] = df_out[col].apply(
                lambda x: round(x * 100, 2) if pd.notna(x) else x
            )

    # Sort summary by verdict order
    df_out["_order"] = df_out["verdict"].apply(
        lambda v: VERDICT_ORDER.index(v) if v in VERDICT_ORDER else 99
    )
    df_summary = df_out.sort_values(
        ["_order", "margin_of_safety_pct"], ascending=[True, False]
    ).drop(columns=["_order"])

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── 1. Summary sheet ──────────────────────────────────────────────────
        df_summary.to_excel(writer, sheet_name="📊 Summary", index=False)
        _style_sheet(writer, "📊 Summary", df_summary)

        # ── 2. Per-verdict sheets ─────────────────────────────────────────────
        for verdict in VERDICT_ORDER:
            subset = df_out[df_out["verdict"] == verdict].drop(columns=["_order"], errors="ignore")
            subset = subset.sort_values("margin_of_safety_pct", ascending=False)
            if subset.empty:
                continue
            sheet_name = verdict[:31]
            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer, sheet_name, subset)

        # ── 3. Monthly Returns sheet (backtest only) ──────────────────────────
        if is_backtest and avail_ret:
            _write_monthly_returns_sheet(writer, df, scan_date)

        # ── 4. Backtest Accuracy sheet ────────────────────────────────────────
        if is_backtest:
            _write_accuracy_sheet(writer, df)

        # ── 5. Metadata ───────────────────────────────────────────────────────
        _write_metadata_sheet(writer, df, scan_date, is_backtest)

    logger.info(f"Excel report saved → {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════════════════════
# MONTHLY RETURNS SHEET
# ══════════════════════════════════════════════════════════════════════════════

def _write_monthly_returns_sheet(writer, df, scan_date):
    """
    Creates a sheet with one row per stock showing:
      - ticker, name, sector, scan_price, verdict, MoS%
      - return at each interval: +30d, +60d, +90d, ..., +365d
    Colour-coded: green=positive, red=negative.
    """
    ret_cols = [c for c in RETURN_COLS if c in df.columns and df[c].notna().any()]
    if not ret_cols:
        return

    cols = ["ticker", "name", "sector", "price", "verdict",
            "margin_of_safety_pct", "quality_score"] + ret_cols

    avail = [c for c in cols if c in df.columns]
    sheet_df = df[avail].copy()
    sheet_df = sheet_df.sort_values("verdict").reset_index(drop=True)

    # Rename return columns for clarity
    rename = {f"return_{d}d": f"+{d}d %" for d in RETURN_INTERVALS}
    sheet_df.rename(columns=rename, inplace=True)

    sheet_df.to_excel(writer, sheet_name="📅 Monthly Returns", index=False)

    # Style
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.sheets["📅 Monthly Returns"]

        # Header
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        green_fill  = PatternFill("solid", fgColor="C6EFCE")
        red_fill    = PatternFill("solid", fgColor="FFC7CE")
        light_green = PatternFill("solid", fgColor="E2EFDA")
        light_red   = PatternFill("solid", fgColor="FFE0E0")

        # Find return column indices
        ret_col_names = [f"+{d}d %" for d in RETURN_INTERVALS]
        ret_col_indices = []
        for i, cell in enumerate(ws[1], 1):
            if cell.value in ret_col_names:
                ret_col_indices.append(i)

        for row in ws.iter_rows(min_row=2):
            for i, cell in enumerate(row, 1):
                if i in ret_col_indices and cell.value is not None:
                    try:
                        v = float(cell.value)
                        cell.fill = green_fill if v >= 10 else \
                                    light_green if v > 0  else \
                                    light_red   if v > -10 else red_fill
                    except Exception:
                        pass

        # Summary rows at bottom (avg per verdict)
        ws.append([])
        ws.append(["AVERAGE RETURN BY VERDICT"] + [""] * (len(avail) - 1))

        for verdict in VERDICT_ORDER:
            subset = df[df["verdict"] == verdict]
            if subset.empty:
                continue
            summary_row = [verdict, "", "", "", "", len(subset), ""]
            for d in RETURN_INTERVALS:
                col = f"return_{d}d"
                if col in subset.columns:
                    vals = subset[col].dropna()
                    summary_row.append(round(vals.mean(), 2) if len(vals) else "")
                else:
                    summary_row.append("")
            ws.append(summary_row[:len(avail)])

        # Column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 20)

        ws.freeze_panes = "A2"

    except Exception as e:
        logger.debug(f"Monthly returns styling error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# BACKTEST ACCURACY SHEET
# ══════════════════════════════════════════════════════════════════════════════

def _write_accuracy_sheet(writer, df):
    """
    Summary table: for each verdict × each time interval,
    shows Avg Return %, Median Return %, % Positive, Count.
    """
    avail_intervals = [d for d in RETURN_INTERVALS
                       if f"return_{d}d" in df.columns
                       and df[f"return_{d}d"].notna().any()]
    if not avail_intervals:
        return

    rows = []
    for verdict in VERDICT_ORDER:
        subset = df[df["verdict"] == verdict]
        if subset.empty:
            continue
        row = {"Verdict": verdict, "Total Stocks": len(subset)}
        for d in avail_intervals:
            col  = f"return_{d}d"
            vals = subset[col].dropna()
            if len(vals) == 0:
                row[f"+{d}d Avg%"]     = ""
                row[f"+{d}d %Positive"] = ""
            else:
                row[f"+{d}d Avg%"]      = round(vals.mean(), 2)
                row[f"+{d}d %Positive"] = round((vals > 0).mean() * 100, 1)
        rows.append(row)

    acc_df = pd.DataFrame(rows)
    acc_df.to_excel(writer, sheet_name="📈 Backtest Accuracy", index=False)
    _style_sheet(writer, "📈 Backtest Accuracy", acc_df)


# ══════════════════════════════════════════════════════════════════════════════
# METADATA SHEET
# ══════════════════════════════════════════════════════════════════════════════

def _write_metadata_sheet(writer, df, scan_date, is_backtest):
    scan_date_str = str(scan_date) if scan_date else "Today (live)"

    def _count(v):
        return len(df[df["verdict"].str.contains(v, na=False)])

    meta = pd.DataFrame({
        "Parameter": [
            "Scan Date", "Scan Time", "Mode",
            "Market Cap Min (Rs cr)", "Market Cap Max (Rs cr)",
            "Total Stocks Scanned",
            "Strong Buy", "Buy", "Hold", "Avoid", "Strong Avoid", "Insufficient Data",
            "", "Valuation Model Weights",
            "  Relative Valuation", "  Graham Number",
            "  DCF (Discounted Cash Flow)", "  Analyst Consensus",
            "", "Financial Parameters",
            "  Risk-Free Rate (India 10yr G-Sec)", "  Equity Risk Premium",
            "  Default WACC", "  Terminal Growth Rate",
            "", "Return Intervals (backtest)",
            "  Intervals tracked",
            "", "Results Season Grey Zones",
            "  Q1 grey zone", "  Q2 grey zone",
            "  Q3 grey zone", "  Q4 grey zone",
        ],
        "Value": [
            scan_date_str,
            datetime.now().strftime("%H:%M:%S"),
            "BACKTEST" if is_backtest else "LIVE",
            df["market_cap_cr"].min() if len(df) else "",
            df["market_cap_cr"].max() if len(df) else "",
            len(df),
            _count("Strong Buy"), _count("Buy"), _count("Hold"),
            _count("Avoid"), _count("Strong Avoid"), _count("Insufficient"),
            "",
            "",
            "35%", "20%", "25%", "20%",
            "",
            "",
            "7.2%", "5.5%", "12%", "3%",
            "",
            "",
            "+30d, +60d, +90d, +120d, +150d, +180d, +210d, +240d, +270d, +300d, +330d, +365d",
            "",
            "",
            "Oct 1-14  (Q1 Apr-Jun results filing period)",
            "Jan 1-14  (Q2 Jul-Sep results filing period)",
            "Apr 1-14  (Q3 Oct-Dec results filing period)",
            "Jul 1-14  (Q4 Jan-Mar results filing period)",
        ]
    })
    meta.to_excel(writer, sheet_name="ℹ️ Metadata", index=False)
    _style_sheet(writer, "ℹ️ Metadata", meta)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET STYLING
# ══════════════════════════════════════════════════════════════════════════════

def _style_sheet(writer, sheet_name, df):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment

        ws = writer.sheets[sheet_name]

        # Header
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Verdict colour fill
        verdict_fills = {
            "⭐ Strong Buy":   PatternFill("solid", fgColor="C6EFCE"),
            "✅ Buy":          PatternFill("solid", fgColor="E2EFDA"),
            "🟡 Hold":         PatternFill("solid", fgColor="FFEB9C"),
            "🔴 Avoid":        PatternFill("solid", fgColor="FFC7CE"),
            "🚫 Strong Avoid": PatternFill("solid", fgColor="FF9999"),
        }

        verdict_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "verdict":
                verdict_col = i
                break

        if verdict_col:
            for row in ws.iter_rows(min_row=2):
                v    = str(row[verdict_col - 1].value or "")
                fill = verdict_fills.get(v)
                if fill:
                    for cell in row:
                        cell.fill = fill

        # Column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 28)

        ws.freeze_panes = "A2"

    except Exception as e:
        logger.debug(f"Styling error on sheet '{sheet_name}': {e}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY
# ══════════════════════════════════════════════════════════════════════════════

def generate_report(df, output_dir="output", scan_date=None, is_backtest=False):
    """Prints console summary and saves the Excel workbook."""
    print_summary(df)
    return save_excel(df, output_dir, scan_date=scan_date, is_backtest=is_backtest)
